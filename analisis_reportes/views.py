from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q, Avg
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import json
import pandas as pd
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.graphics import renderPDF
from io import BytesIO
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import base64
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from django.utils.text import slugify

from core.decorators import fast_access_pin_verified
from .models import Reporte, ConfiguracionReporte
from cuentas.models import Cuenta, SubCuenta, TransferenciaSubCuenta
from gestion_financiera_basica.models import Movimiento, MetaAhorro

@login_required
@fast_access_pin_verified
def reports(request):
    """Vista principal de reportes con dashboard interactivo"""
    # Configuración del usuario
    config, created = ConfiguracionReporte.objects.get_or_create(
        id_usuario=request.user,
        defaults={'periodo_default': 'mes_actual'}
    )
    
    # Obtener parámetros de filtro
    periodo = request.GET.get('periodo', config.periodo_default)
    
    # Manejar fechas personalizadas
    if periodo == 'personalizado':
        fecha_inicio_str = request.GET.get('fecha_inicio')
        fecha_fin_str = request.GET.get('fecha_fin')
        
        if fecha_inicio_str and fecha_fin_str:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            except ValueError:
                fecha_inicio, fecha_fin = get_periodo_fechas('mes_actual')
        else:
            fecha_inicio, fecha_fin = get_periodo_fechas('mes_actual')
    else:
        fecha_inicio, fecha_fin = get_periodo_fechas(periodo)
    
    # Si es una request AJAX para actualizar filtros
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            stats = calcular_estadisticas_generales(request.user, fecha_inicio, fecha_fin)
            gastos_categoria = get_gastos_por_categoria(request.user, fecha_inicio, fecha_fin)
            ingresos_egresos = get_ingresos_vs_egresos(request.user, fecha_inicio, fecha_fin)
            subcuentas_data = get_estadisticas_subcuentas(request.user)
            flujo_mensual = get_flujo_mensual(request.user, fecha_inicio, fecha_fin)
            
            return JsonResponse({
                'success': True,
                'stats': stats,
                'gastos_categoria': gastos_categoria,
                'ingresos_egresos': ingresos_egresos,
                'subcuentas_data': subcuentas_data,
                'flujo_mensual': flujo_mensual,
                'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'),
                'fecha_fin': fecha_fin.strftime('%Y-%m-%d'),
                'periodo': periodo,
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    # Obtener cuentas del usuario
    cuentas = Cuenta.objects.filter(id_usuario=request.user)
    
    # Calcular estadísticas generales
    stats = calcular_estadisticas_generales(request.user, fecha_inicio, fecha_fin)
    
    # Datos para gráficos
    gastos_categoria = get_gastos_por_categoria(request.user, fecha_inicio, fecha_fin)
    ingresos_egresos = get_ingresos_vs_egresos(request.user, fecha_inicio, fecha_fin)
    subcuentas_data = get_estadisticas_subcuentas(request.user)
    flujo_mensual = get_flujo_mensual(request.user, fecha_inicio, fecha_fin)
    
    # Reportes recientes
    reportes_recientes = Reporte.objects.filter(id_usuario=request.user).order_by('-fecha_creacion')[:5]
    
    context = {
        'stats': stats,
        'gastos_categoria': json.dumps(gastos_categoria),
        'ingresos_egresos': json.dumps(ingresos_egresos),
        'subcuentas_data': json.dumps(subcuentas_data),
        'flujo_mensual': json.dumps(flujo_mensual),
        'reportes_recientes': reportes_recientes,
        'periodo_actual': periodo,
        'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'),
        'fecha_fin': fecha_fin.strftime('%Y-%m-%d'),
        'config': config,
        'cuentas': cuentas,
    }
    
    return render(request, 'analisis_reportes/reports.html', context)

@login_required
@fast_access_pin_verified
def generar_reporte(request):
    """Vista para generar un reporte específico"""
    if request.method == 'POST':
        tipo_reporte = request.POST.get('tipo_reporte')
        titulo = request.POST.get('titulo', f'Reporte {tipo_reporte}')
        fecha_inicio = datetime.strptime(request.POST.get('fecha_inicio'), '%Y-%m-%d')
        fecha_fin = datetime.strptime(request.POST.get('fecha_fin'), '%Y-%m-%d')
        
        # Generar datos según el tipo de reporte
        datos_reporte = {}
        
        if tipo_reporte == 'gastos_categoria':
            datos_reporte = get_gastos_por_categoria(request.user, fecha_inicio, fecha_fin)
        elif tipo_reporte == 'ingresos_egresos':
            datos_reporte = get_ingresos_vs_egresos(request.user, fecha_inicio, fecha_fin)
        elif tipo_reporte == 'subcuentas_analisis':
            datos_reporte = get_estadisticas_subcuentas(request.user)
        elif tipo_reporte == 'balance_general':
            datos_reporte = get_balance_general(request.user, fecha_inicio, fecha_fin)
        elif tipo_reporte == 'flujo_efectivo':
            datos_reporte = get_flujo_mensual(request.user, fecha_inicio, fecha_fin)
        
        # Crear el reporte
        reporte = Reporte.objects.create(
            tipo_reporte=tipo_reporte,
            titulo=titulo,
            descripcion=request.POST.get('descripcion', ''),
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            id_usuario=request.user,
            datos_json=json.dumps(datos_reporte, default=str)
        )
        
        messages.success(request, f'Reporte "{titulo}" generado exitosamente.')
        
        # Si se solicita exportación inmediata
        formato_export = request.POST.get('formato_export')
        if formato_export:
            return exportar_reporte(request, reporte.id, formato_export)
        
        return redirect('analisis_reportes:ver_reporte', reporte_id=reporte.id)
    
    return redirect('analisis_reportes:reports')

@login_required
@fast_access_pin_verified
def ver_reporte(request, reporte_id):
    """Vista para ver un reporte específico"""
    reporte = get_object_or_404(Reporte, id=reporte_id, id_usuario=request.user)
    datos = reporte.get_datos()
    
    # Procesar datos para el template
    datos_procesados = procesar_datos_para_template(reporte.tipo_reporte, datos)
    
    context = {
        'reporte': reporte,
        'datos': datos,
        'datos_procesados': datos_procesados,
        'datos_json': json.dumps(datos),
    }
    
    return render(request, 'analisis_reportes/ver_reporte.html', context)

@login_required
@fast_access_pin_verified  
def exportar_reporte(request, reporte_id, formato):
    """Vista para exportar reportes en diferentes formatos"""
    reporte = get_object_or_404(Reporte, id=reporte_id, id_usuario=request.user)
    datos = reporte.get_datos()
    
    if formato == 'pdf':
        return exportar_pdf(reporte, datos)
    elif formato == 'excel':
        return exportar_excel(reporte, datos)
    elif formato == 'csv':
        return exportar_csv(reporte, datos)
    else:
        messages.error(request, 'Formato de exportación no válido.')
        return redirect('analisis_reportes:ver_reporte', reporte_id=reporte_id)

@login_required
@fast_access_pin_verified
def api_datos_grafico(request):
    """API para obtener datos de gráficos via AJAX"""
    tipo = request.GET.get('tipo')
    periodo = request.GET.get('periodo', 'mes_actual')
    fecha_inicio, fecha_fin = get_periodo_fechas(periodo)
    
    if tipo == 'gastos_categoria':
        datos = get_gastos_por_categoria(request.user, fecha_inicio, fecha_fin)
    elif tipo == 'ingresos_egresos':
        datos = get_ingresos_vs_egresos(request.user, fecha_inicio, fecha_fin)
    elif tipo == 'subcuentas':
        datos = get_estadisticas_subcuentas(request.user)
    elif tipo == 'flujo_mensual':
        datos = get_flujo_mensual(request.user, fecha_inicio, fecha_fin)
    else:
        datos = {}
    
    return JsonResponse(datos)

# Funciones auxiliares para cálculos

def get_periodo_fechas(periodo):
    """Convierte un período en fechas de inicio y fin"""
    hoy = timezone.now().date()
    
    if periodo == 'semana_actual':
        inicio = hoy - timedelta(days=hoy.weekday())
        fin = inicio + timedelta(days=6)
    elif periodo == 'mes_actual':
        inicio = hoy.replace(day=1)
        fin = (inicio + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    elif periodo == 'trimestre_actual':
        mes_inicio = ((hoy.month - 1) // 3) * 3 + 1
        inicio = hoy.replace(month=mes_inicio, day=1)
        fin = (inicio + timedelta(days=92)).replace(day=1) - timedelta(days=1)
    elif periodo == 'año_actual':
        inicio = hoy.replace(month=1, day=1)
        fin = hoy.replace(month=12, day=31)
    elif periodo == 'ultimos_30_dias':
        fin = hoy
        inicio = hoy - timedelta(days=30)
    elif periodo == 'ultimos_90_dias':
        fin = hoy
        inicio = hoy - timedelta(days=90)
    else:
        # Por defecto, mes actual
        inicio = hoy.replace(day=1)
        fin = (inicio + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    return inicio, fin

def calcular_estadisticas_generales(usuario, fecha_inicio, fecha_fin):
    """Calcula estadísticas generales del usuario"""
    cuentas = Cuenta.objects.filter(id_usuario=usuario)
    
    # Balance total
    balance_total = sum([cuenta.saldo_cuenta for cuenta in cuentas])
    
    # Total en subcuentas (incluir independientes)
    total_subcuentas = SubCuenta.objects.filter(
        Q(id_cuenta__id_usuario=usuario) | Q(propietario=usuario), 
        activa=True
    ).aggregate(total=Sum('saldo'))['total'] or 0
    
    # Transacciones del período
    transacciones_periodo = Movimiento.objects.filter(
        id_cuenta__id_usuario=usuario,
        fecha_movimiento__range=[fecha_inicio, fecha_fin]
    )
    
    ingresos_periodo = transacciones_periodo.filter(
        tipo='ingreso'
    ).aggregate(total=Sum('monto'))['total'] or 0
    
    gastos_periodo = transacciones_periodo.filter(
        tipo='egreso'
    ).aggregate(total=Sum('monto'))['total'] or 0
    
    # Número de cuentas y subcuentas (incluir independientes)
    num_cuentas = cuentas.count()
    num_subcuentas = SubCuenta.objects.filter(
        Q(id_cuenta__id_usuario=usuario) | Q(propietario=usuario), 
        activa=True
    ).count()
    
    # Promedio de transacciones
    promedio_transaccion = transacciones_periodo.aggregate(
        promedio=Avg('monto')
    )['promedio'] or 0
    
    # Top gastos
    top_gastos = transacciones_periodo.filter(tipo='egreso').order_by('-monto')[:5]
    
    # Información de subcuentas
    subcuentas_info = SubCuenta.objects.filter(
        Q(id_cuenta__id_usuario=usuario) | Q(propietario=usuario), 
        activa=True
    )[:5]
    
    # Metas de ahorro con progreso
    from gestion_financiera_basica.models import MetaAhorro
    metas_progreso = MetaAhorro.objects.filter(id_usuario=usuario)[:3]
    for meta in metas_progreso:
        if meta.monto_objetivo > 0:
            meta.porcentaje_progreso = (float(meta.monto_ahorrado()) / float(meta.monto_objetivo)) * 100
        else:
            meta.porcentaje_progreso = 0
    
    return {
        'balance_total': float(balance_total),
        'total_subcuentas': float(total_subcuentas),
        'total_ingresos': float(ingresos_periodo),  # Cambio de nombre para el template
        'total_egresos': float(gastos_periodo),      # Cambio de nombre para el template
        'ahorro_neto': float(ingresos_periodo - gastos_periodo),  # Nuevo campo
        'num_cuentas': num_cuentas,
        'num_subcuentas': num_subcuentas,
        'promedio_transaccion': float(promedio_transaccion),
        'total_transacciones': transacciones_periodo.count(),
        'top_gastos': top_gastos,
        'subcuentas_info': subcuentas_info,
        'metas_progreso': metas_progreso,
    }

def get_gastos_por_categoria(usuario, fecha_inicio, fecha_fin):
    """Obtiene gastos agrupados por categoría de movimientos"""
    # Obtener gastos agrupados por nombre (categoría)
    gastos_por_categoria = Movimiento.objects.filter(
        id_cuenta__id_usuario=usuario,
        tipo='egreso',
        fecha_movimiento__range=[fecha_inicio, fecha_fin]
    ).values('nombre').annotate(
        total=Sum('monto'),
        count=Count('id')
    ).order_by('-total')[:10]
    
    labels = []
    values = []
    
    if gastos_por_categoria.exists():
        for gasto in gastos_por_categoria:
            labels.append(gasto['nombre'] or 'Sin categoría')
            values.append(float(gasto['total']))
    else:
        # Datos por defecto si no hay gastos
        labels = ['Sin gastos registrados']
        values = [0]
    
    return {
        'labels': labels,
        'values': values,
    }

def get_ingresos_vs_egresos(usuario, fecha_inicio, fecha_fin):
    """Obtiene comparación de ingresos vs egresos por mes"""
    from datetime import datetime, timedelta
    import calendar
    
    # Obtener datos mes por mes en el rango
    fecha_actual = fecha_inicio.replace(day=1)
    labels = []
    ingresos = []
    gastos = []
    
    while fecha_actual <= fecha_fin:
        # Último día del mes
        ultimo_dia = calendar.monthrange(fecha_actual.year, fecha_actual.month)[1]
        fin_mes = fecha_actual.replace(day=ultimo_dia)
        
        # Transacciones del mes
        transacciones_mes = Movimiento.objects.filter(
            id_cuenta__id_usuario=usuario,
            fecha_movimiento__range=[fecha_actual, fin_mes]
        )
        
        ingresos_mes = transacciones_mes.filter(tipo='ingreso').aggregate(
            total=Sum('monto'))['total'] or 0
        gastos_mes = transacciones_mes.filter(tipo='egreso').aggregate(
            total=Sum('monto'))['total'] or 0
        
        labels.append(fecha_actual.strftime('%b %Y'))
        ingresos.append(float(ingresos_mes))
        gastos.append(float(gastos_mes))
        
        # Siguiente mes
        if fecha_actual.month == 12:
            fecha_actual = fecha_actual.replace(year=fecha_actual.year + 1, month=1)
        else:
            fecha_actual = fecha_actual.replace(month=fecha_actual.month + 1)
    
    return {
        'labels': labels,
        'ingresos': ingresos,
        'gastos': gastos,
    }

def get_estadisticas_subcuentas(usuario):
    """Obtiene estadísticas de subcuentas"""
    # Incluir tanto subcuentas vinculadas como independientes
    subcuentas = SubCuenta.objects.filter(
        Q(id_cuenta__id_usuario=usuario) | Q(propietario=usuario),
        activa=True
    ).select_related('id_cuenta')
    
    if not subcuentas.exists():
        return {
            'labels': ['Sin subcuentas'],
            'saldos': [0],
            'cantidades': [0],
            'tipos_detalle': [],
        }
    
    # Agrupar por tipo de subcuenta
    tipos_subcuentas = {}
    tipos_detalle = []
    
    for subcuenta in subcuentas:
        tipo = subcuenta.tipo or 'otros'
        tipo_display = subcuenta.get_tipo_display() or tipo.title()
        
        # Para la visualización, usar el display name
        if tipo_display not in tipos_subcuentas:
            tipos_subcuentas[tipo_display] = {
                'saldo_total': 0,
                'cantidad': 0,
                'es_negocio': subcuenta.es_negocio,
                'subcuentas': []
            }
        
        tipos_subcuentas[tipo_display]['saldo_total'] += float(subcuenta.saldo)
        tipos_subcuentas[tipo_display]['cantidad'] += 1
        tipos_subcuentas[tipo_display]['subcuentas'].append({
            'nombre': subcuenta.nombre,
            'saldo': float(subcuenta.saldo),
            'es_independiente': subcuenta.es_independiente(),
            'color': subcuenta.color
        })
        
        # Agregar detalles para el template
        tipos_detalle.append({
            'nombre': subcuenta.nombre,
            'tipo': tipo_display,
            'saldo': float(subcuenta.saldo),
            'es_negocio': subcuenta.es_negocio,
            'es_independiente': subcuenta.es_independiente(),
            'color': subcuenta.color
        })
    
    labels = list(tipos_subcuentas.keys())
    saldos = [data['saldo_total'] for data in tipos_subcuentas.values()]
    cantidades = [data['cantidad'] for data in tipos_subcuentas.values()]
    
    return {
        'labels': labels,
        'saldos': saldos,
        'cantidades': cantidades,
        'tipos_detalle': tipos_detalle,
        'total_subcuentas': len(subcuentas),
        'total_saldo': sum(saldos),
        'tipos_subcuentas': tipos_subcuentas,
    }

def get_flujo_mensual(usuario, fecha_inicio, fecha_fin):
    """Obtiene flujo de efectivo mensual"""
    from datetime import datetime, timedelta
    import calendar
    
    labels = []
    values = []
    
    fecha_actual = fecha_inicio.replace(day=1)
    
    while fecha_actual <= fecha_fin:
        # Último día del mes
        ultimo_dia = calendar.monthrange(fecha_actual.year, fecha_actual.month)[1]
        fin_mes = fecha_actual.replace(day=ultimo_dia)
        
        # Transacciones del mes
        transacciones_mes = Movimiento.objects.filter(
            id_cuenta__id_usuario=usuario,
            fecha_movimiento__range=[fecha_actual, fin_mes]
        )
        
        ingresos_mes = transacciones_mes.filter(tipo='ingreso').aggregate(
            total=Sum('monto'))['total'] or 0
        gastos_mes = transacciones_mes.filter(tipo='egreso').aggregate(
            total=Sum('monto'))['total'] or 0
        
        flujo_neto = float(ingresos_mes - gastos_mes)
        
        labels.append(fecha_actual.strftime('%b %Y'))
        values.append(flujo_neto)
        
        # Siguiente mes
        if fecha_actual.month == 12:
            fecha_actual = fecha_actual.replace(year=fecha_actual.year + 1, month=1)
        else:
            fecha_actual = fecha_actual.replace(month=fecha_actual.month + 1)
    
    # Si no hay datos, mostrar al menos el mes actual
    if not labels:
        labels = [datetime.now().strftime('%b %Y')]
        values = [0]
    
    return {
        'labels': labels,
        'values': values,
    }

def get_balance_general(usuario, fecha_inicio, fecha_fin):
    """Obtiene balance general del período"""
    cuentas = Cuenta.objects.filter(id_usuario=usuario)
    
    balance_data = []
    for cuenta in cuentas:
        subcuentas = SubCuenta.objects.filter(id_cuenta=cuenta, activa=True)
        total_subcuentas = sum([sc.saldo for sc in subcuentas])
        
        balance_data.append({
            'cuenta': cuenta.nombre,
            'saldo_principal': float(cuenta.saldo_cuenta),
            'saldo_subcuentas': float(total_subcuentas),
            'saldo_total': float(cuenta.saldo_cuenta + total_subcuentas),
            'subcuentas': [{
                'nombre': sc.nombre,
                'tipo': sc.tipo,
                'saldo': float(sc.saldo)
            } for sc in subcuentas]
        })
    
    return balance_data

def exportar_pdf(reporte, datos):
    """Exporta reporte a PDF con formato ultra profesional y visualmente atractivo"""
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.graphics import renderPDF
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=50, leftMargin=50,
        topMargin=80, bottomMargin=80
    )
    styles = getSampleStyleSheet()
    story = []
    
    # Estilos personalizados ultra profesionales
    title_style = ParagraphStyle(
        'UltraTitle',
        parent=styles['Heading1'],
        fontSize=28,
        textColor=colors.HexColor('#2C3E50'),
        alignment=TA_CENTER,
        spaceAfter=35,
        fontName='Helvetica-Bold',
        leading=32
    )
    
    brand_style = ParagraphStyle(
        'BrandStyle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#6C5CE7'),
        alignment=TA_CENTER,
        spaceBefore=5,
        spaceAfter=25,
        fontName='Helvetica-Bold'
    )
    
    header_style = ParagraphStyle(
        'ProfessionalHeader',
        parent=styles['Heading2'],
        fontSize=18,
        textColor=colors.HexColor('#34495E'),
        spaceBefore=25,
        spaceAfter=15,
        fontName='Helvetica-Bold',
        borderWidth=0,
        borderPadding=0
    )
    
    subheader_style = ParagraphStyle(
        'SubHeader',
        parent=styles['Heading3'],
        fontSize=14,
        textColor=colors.HexColor('#5D6D7E'),
        spaceBefore=15,
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    
    summary_style = ParagraphStyle(
        'SummaryStyle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#2C3E50'),
        spaceBefore=5,
        spaceAfter=5,
        fontName='Helvetica',
        leading=14
    )
    
    # ENCABEZADO ULTRA PROFESIONAL
    # Crear un rectángulo de fondo para el header
    header_bg = Drawing(500, 60)
    header_bg.add(Rect(0, 0, 500, 60, fillColor=colors.HexColor('#F8F9FA'), strokeColor=None))
    story.append(header_bg)
    story.append(Spacer(1, -50))
    
    # Logo y brand (simulado con texto estilizado)
    story.append(Paragraph("FINGEST", brand_style))
    story.append(Paragraph("REPORTE FINANCIERO PROFESIONAL", title_style))
    
    # Línea decorativa
    line_drawing = Drawing(500, 5)
    line_drawing.add(Rect(0, 2, 500, 1, fillColor=colors.HexColor('#6C5CE7'), strokeColor=None))
    story.append(line_drawing)
    story.append(Spacer(1, 20))
    
    # INFORMACIÓN DEL REPORTE CON DISEÑO PREMIUM
    story.append(Paragraph("INFORMACIÓN DEL REPORTE", header_style))
    
    # Crear tabla de información más elegante
    info_data = [
        ['Tipo de Reporte:', reporte.get_tipo_reporte_display()],
        ['Fecha de Generación:', reporte.fecha_creacion.strftime('%d de %B de %Y a las %H:%M hrs')],
        ['Período Analizado:', f"Del {reporte.fecha_inicio.strftime('%d de %B de %Y')} al {reporte.fecha_fin.strftime('%d de %B de %Y')}"],
        ['Usuario:', str(reporte.id_usuario) if reporte.id_usuario else 'No especificado'],
    ]
    
    if reporte.descripcion:
        info_data.append(['Descripción:', reporte.descripcion])
    
    info_table = Table(info_data, colWidths=[2.8*inch, 4.2*inch])
    info_table.setStyle(TableStyle([
        # Estilo del encabezado
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F4FD')),
        ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#FFFFFF')),
        
        # Bordes y líneas
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
        ('LINEBEFORE', (1, 0), (1, -1), 0.5, colors.HexColor('#BDC3C7')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#34495E')),
        
        # Fuentes y colores
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#34495E')),
        
        # Alineación y espaciado
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('LEFTPADDING', (0, 0), (-1, -1), 15),
        ('RIGHTPADDING', (0, 0), (-1, -1), 15),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 35))
    
    # Contenido específico según tipo de reporte con diseño ultra premium
    if reporte.tipo_reporte == 'gastos_categoria' and 'labels' in datos:
        story.append(Paragraph("ANÁLISIS DETALLADO DE GASTOS POR CATEGORÍA", header_style))
        
        # Resumen ejecutivo
        total = sum(datos['data'])
        total_transacciones = sum(datos.get('counts', []))
        promedio_por_categoria = total / len(datos['labels']) if datos['labels'] else 0
        
        summary_data = [
            ['RESUMEN EJECUTIVO'],
            [f'Gasto Total: ${total:,.2f} MXN'],
            [f'Total de Transacciones: {total_transacciones:,}'],
            [f'Promedio por Categoría: ${promedio_por_categoria:,.2f} MXN'],
            [f'Categorías Analizadas: {len(datos["labels"])}']
        ]
        
        summary_table = Table(summary_data, colWidths=[7*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#EBF5FF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2C3E50')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('FONTSIZE', (0, 1), (-1, -1), 11),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#3498DB')),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 25))
        
        # Tabla principal de gastos con diseño premium
        story.append(Paragraph("DESGLOSE DETALLADO POR CATEGORÍA", subheader_style))
        
        gastos_data = [['CATEGORÍA', 'MONTO (MXN)', 'CANTIDAD', 'PORCENTAJE', 'PROMEDIO']]
        
        for i, (label, monto, count) in enumerate(zip(datos['labels'], datos['data'], datos.get('counts', [0]*len(datos['labels'])))):
            porcentaje = (monto / total * 100) if total > 0 else 0
            promedio_item = monto / count if count > 0 else 0
            
            gastos_data.append([
                label,
                f"${monto:,.2f}",
                f"{count:,}",
                f"{porcentaje:.1f}%",
                f"${promedio_item:,.2f}"
            ])
        
        # Fila de total con estilo especial
        gastos_data.append([
            'TOTAL GENERAL', 
            f"${total:,.2f}", 
            f"{sum(datos.get('counts', [])):,}", 
            '100.0%',
            f"${total/len(datos['labels']) if datos['labels'] else 0:,.2f}"
        ])
        
        gastos_table = Table(gastos_data, colWidths=[2.2*inch, 1.3*inch, 1*inch, 1*inch, 1.2*inch])
        gastos_table.setStyle(TableStyle([
            # Encabezado principal
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Filas de datos
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 10),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            
            # Fila de total
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F8F5')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#27AE60')),
            
            # Alternancia de colores en filas
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F9FA')]),
            
            # Bordes y líneas
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#2C3E50')),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#27AE60')),
            
            # Espaciado
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        story.append(gastos_table)
        
        # Insights automáticos
        story.append(Spacer(1, 25))
        story.append(Paragraph("INSIGHTS AUTOMÁTICOS", subheader_style))
        
        # Encontrar la categoría con mayor gasto
        max_index = datos['data'].index(max(datos['data'])) if datos['data'] else 0
        categoria_mayor = datos['labels'][max_index] if max_index < len(datos['labels']) else 'N/A'
        porcentaje_mayor = (max(datos['data']) / total * 100) if total > 0 else 0
        
        insights = [
            f"• La categoría '{categoria_mayor}' representa el mayor gasto con {porcentaje_mayor:.1f}% del total",
            f"• Promedio de gasto por transacción: ${total/total_transacciones if total_transacciones > 0 else 0:,.2f} MXN",
            f"• Distribución: {len([x for x in datos['data'] if x > promedio_por_categoria])} categorías están por encima del promedio"
        ]
        
        for insight in insights:
            story.append(Paragraph(insight, summary_style))
            story.append(Spacer(1, 5))
        
    elif reporte.tipo_reporte == 'ingresos_egresos':
        story.append(Paragraph("ANÁLISIS FINANCIERO: INGRESOS VS EGRESOS", header_style))
        
        ingresos = datos['data'][0] if len(datos['data']) > 0 else 0
        egresos = datos['data'][1] if len(datos['data']) > 1 else 0
        balance = ingresos - egresos
        tasa_ahorro = (balance / ingresos * 100) if ingresos > 0 else 0
        
        # Dashboard financiero visual
        dashboard_data = [
            ['DASHBOARD FINANCIERO EJECUTIVO'],
            [f'Total de Ingresos: ${ingresos:,.2f} MXN'],
            [f'Total de Egresos: ${egresos:,.2f} MXN'],
            [f'Balance Neto: ${balance:,.2f} MXN'],
            [f'Tasa de Ahorro: {tasa_ahorro:.1f}%'],
            [f'Ratio de Gastos: {(egresos/ingresos*100) if ingresos > 0 else 0:.1f}%']
        ]
        
        # Color del balance
        balance_color = colors.HexColor('#27AE60') if balance >= 0 else colors.HexColor('#E74C3C')
        
        dashboard_table = Table(dashboard_data, colWidths=[7*inch])
        dashboard_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 1), (0, 2), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 3), (0, 3), balance_color),  # Color del balance
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#34495E')),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        story.append(dashboard_table)
        story.append(Spacer(1, 25))
        
        # Tabla comparativa detallada
        story.append(Paragraph("ANÁLISIS COMPARATIVO DETALLADO", subheader_style))
        
        ie_data = [
            ['CONCEPTO', 'MONTO (MXN)', 'PORCENTAJE', 'EVALUACIÓN'],
            ['Ingresos Totales', f"${ingresos:,.2f}", '100.0%', 'Base de cálculo'],
            ['Egresos Totales', f"${egresos:,.2f}", f"{(egresos/ingresos*100):.1f}%" if ingresos > 0 else '0.0%', 
             'Alto' if egresos/ingresos > 0.8 else 'Controlado' if egresos/ingresos <= 0.6 else 'Moderado'],
            ['Balance Final', f"${balance:,.2f}", f"{(balance/ingresos*100):.1f}%" if ingresos > 0 else '0.0%',
             'Excelente' if balance > 0 else 'Déficit'],
        ]
        
        ie_table = Table(ie_data, colWidths=[2*inch, 1.8*inch, 1.5*inch, 1.7*inch])
        ie_table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Filas de datos
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 10),
            ('ALIGN', (1, 1), (2, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('ALIGN', (3, 1), (3, -1), 'CENTER'),
            
            # Colores especiales para balance
            ('TEXTCOLOR', (0, -1), (-1, -1), balance_color),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            
            # Alternancia de colores
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#2C3E50')),
            
            # Espaciado
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        
        story.append(ie_table)
        
        # Recomendaciones automáticas
        story.append(Spacer(1, 25))
        story.append(Paragraph("RECOMENDACIONES FINANCIERAS", subheader_style))
        
        recomendaciones = []
        if balance < 0:
            recomendaciones.append("URGENTE: Tienes un déficit financiero. Revisa tus gastos y busca reducir egresos.")
            recomendaciones.append("Elabora un presupuesto detallado para identificar gastos innecesarios.")
        elif tasa_ahorro < 10:
            recomendaciones.append("Tu tasa de ahorro es baja. Considera incrementar tus ingresos o reducir gastos.")
            recomendaciones.append("Meta recomendada: Ahorra al menos el 15-20% de tus ingresos.")
        else:
            recomendaciones.append("¡Excelente! Mantienes un balance positivo.")
            recomendaciones.append("Considera invertir tu excedente para hacer crecer tu patrimonio.")
        
        for rec in recomendaciones:
            story.append(Paragraph(f"• {rec}", summary_style))
            story.append(Spacer(1, 5))
        
    elif reporte.tipo_reporte == 'subcuentas_analisis' and 'labels' in datos:
        story.append(Paragraph("ANÁLISIS INTEGRAL DE SUBCUENTAS", header_style))
        
        # Resumen ejecutivo de subcuentas
        total_saldo = sum(datos['saldos'])
        total_cantidad = sum(datos['cantidades'])
        promedio_general = total_saldo / total_cantidad if total_cantidad > 0 else 0
        
        resumen_data = [
            ['RESUMEN EJECUTIVO DE SUBCUENTAS'],
            [f'Patrimonio Total: ${total_saldo:,.2f} MXN'],
            [f'Subcuentas Activas: {total_cantidad}'],
            [f'Saldo Promedio: ${promedio_general:,.2f} MXN'],
            [f'Tipos de Cuenta: {len(datos["labels"])}'],
            [f'Diversificación: {"Alta" if len(datos["labels"]) >= 3 else "Media" if len(datos["labels"]) == 2 else "Baja"}']
        ]
        
        resumen_table = Table(resumen_data, colWidths=[7*inch])
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8E44AD')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F4F2FF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2C3E50')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('FONTSIZE', (0, 1), (-1, -1), 11),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#8E44AD')),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        
        story.append(resumen_table)
        story.append(Spacer(1, 25))
        
        # Tabla detallada de subcuentas
        story.append(Paragraph("DESGLOSE DETALLADO POR TIPO DE SUBCUENTA", subheader_style))
        
        sc_data = [['TIPO DE SUBCUENTA', 'SALDO TOTAL', 'CANTIDAD', '% DEL TOTAL', 'PROMEDIO']]
        
        for i, label in enumerate(datos['labels']):
            saldo = datos['saldos'][i] if i < len(datos['saldos']) else 0
            cantidad = datos['cantidades'][i] if i < len(datos['cantidades']) else 0
            promedio = saldo / cantidad if cantidad > 0 else 0
            porcentaje = (saldo / total_saldo * 100) if total_saldo > 0 else 0
            
            sc_data.append([
                label,
                f"${saldo:,.2f}",
                str(cantidad),
                f"{porcentaje:.1f}%",
                f"${promedio:,.2f}"
            ])
        
        # Fila de totales
        sc_data.append([
            'TOTAL PATRIMONIAL', 
            f"${total_saldo:,.2f}", 
            str(total_cantidad), 
            '100.0%',
            f"${promedio_general:,.2f}"
        ])
        
        sc_table = Table(sc_data, colWidths=[2.2*inch, 1.4*inch, 1*inch, 1*inch, 1.2*inch])
        sc_table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8E44AD')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Filas de datos
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 10),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            
            # Fila de total
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F5E8')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#27AE60')),
            
            # Alternancia de colores
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F9FA')]),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#8E44AD')),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#27AE60')),
            
            # Espaciado
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        
        story.append(sc_table)
        
        # Análisis y recomendaciones
        story.append(Spacer(1, 25))
        story.append(Paragraph("ANÁLISIS Y RECOMENDACIONES ESTRATÉGICAS", subheader_style))
        
        # Encontrar la subcuenta principal
        if datos['saldos']:
            max_saldo_idx = datos['saldos'].index(max(datos['saldos']))
            subcuenta_principal = datos['labels'][max_saldo_idx]
            porcentaje_principal = (max(datos['saldos']) / total_saldo * 100) if total_saldo > 0 else 0
            
            analisis = [
                f"• Subcuenta Principal: '{subcuenta_principal}' concentra {porcentaje_principal:.1f}% del patrimonio",
                f"• Distribución: {'Balanceada' if porcentaje_principal < 60 else 'Concentrada en una subcuenta'}",
                f"• Diversificación: {'Excelente' if len(datos['labels']) >= 4 else 'Buena' if len(datos['labels']) >= 2 else 'Limitada'}",
                f"• Saldo promedio por subcuenta: ${promedio_general:,.2f} MXN"
            ]
            
            if porcentaje_principal > 70:
                analisis.append("• Recomendación: Considera diversificar más tu patrimonio en diferentes subcuentas")
            elif total_cantidad < 3:
                analisis.append("• Sugerencia: Podrías beneficiarte de crear subcuentas adicionales para mejor organización")
            else:
                analisis.append("• Excelente organización financiera con buena distribución de subcuentas")
            
            for insight in analisis:
                story.append(Paragraph(insight, summary_style))
                story.append(Spacer(1, 5))
    
    # PIE DE PÁGINA PROFESIONAL Y BRAND
    story.append(Spacer(1, 40))
    
    # Línea separadora elegante
    separator = Drawing(500, 3)
    separator.add(Rect(0, 1, 500, 1, fillColor=colors.HexColor('#6C5CE7'), strokeColor=None))
    story.append(separator)
    story.append(Spacer(1, 20))
    
    # Información de contacto y legal
    footer_content = [
        ['FinGest - Gestión Financiera Inteligente'],
        ['Soporte: soporte@fingest.com | Tel: +52 (55) 1234-5678'],
        ['www.fingest.com | Descarga nuestra app móvil'],
        ['Este reporte es confidencial y de uso exclusivo del titular'],
        [f'Generado el {datetime.now().strftime("%d de %B de %Y a las %H:%M hrs")}']
    ]
    
    footer_table = Table(footer_content, colWidths=[7*inch])
    footer_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#7F8C8D')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    story.append(footer_table)
    
    # Marca de agua de seguridad
    story.append(Spacer(1, 10))
    security_style = ParagraphStyle(
        'SecurityStyle',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.HexColor('#95A5A6'),
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    story.append(Paragraph("Documento generado con tecnología segura FinGest | ID: " + 
                          f"FG-{reporte.id}-{reporte.fecha_creacion.strftime('%Y%m%d%H%M')}", security_style))
    
    doc.build(story)
    buffer.seek(0)
    
    # Nombre de archivo súper descriptivo
    tipo_clean = reporte.get_tipo_reporte_display().replace(' ', '_').replace('/', '-')
    fecha_str = reporte.fecha_creacion.strftime('%Y%m%d_%H%M')
    filename = f"FinGest_Reporte_{tipo_clean}_{fecha_str}.pdf"
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def exportar_excel(reporte, datos):
    """Exporta reporte a Excel con formato ultra profesional y limpio"""
    from openpyxl import Workbook
    from django.http import HttpResponse
    from datetime import datetime
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Financiero"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Título
    ws['A1'] = f"Reporte Financiero - {reporte.tipo_reporte.replace('_', ' ').title()}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')
    
    # Fecha del reporte
    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A3'] = f"Período: {reporte.fecha_inicio.strftime('%d/%m/%Y')} - {reporte.fecha_fin.strftime('%d/%m/%Y')}"
    
    # Estadísticas generales
    stats = calcular_estadisticas_generales(reporte.id_usuario, reporte.fecha_inicio, reporte.fecha_fin)
    
    ws['A5'] = "RESUMEN GENERAL"
    ws['A5'].font = header_font
    ws['A5'].fill = header_fill
    ws.merge_cells('A5:B5')
    
    row = 6
    resumen_data = [
        ["Balance Total:", f"${stats['balance_total']:,.2f}"],
        ["Total Ingresos:", f"${stats['total_ingresos']:,.2f}"],
        ["Total Gastos:", f"${stats['total_egresos']:,.2f}"],
        ["Ahorro Neto:", f"${stats['ahorro_neto']:,.2f}"],
        ["Total Transacciones:", stats['total_transacciones']],
    ]
    
    for data in resumen_data:
        ws[f'A{row}'] = data[0]
        ws[f'B{row}'] = data[1]
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Transacciones detalladas
    ws['A' + str(row + 2)] = "TRANSACCIONES DETALLADAS"
    ws['A' + str(row + 2)].font = header_font
    ws['A' + str(row + 2)].fill = header_fill
    ws.merge_cells(f'A{row + 2}:E{row + 2}')
    
    # Headers de transacciones
    row += 4
    headers = ['Fecha', 'Descripción', 'Tipo', 'Monto', 'Cuenta']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Datos de transacciones
    transacciones = Movimiento.objects.filter(
        id_cuenta__id_usuario=reporte.id_usuario,
        fecha_movimiento__range=[reporte.fecha_inicio, reporte.fecha_fin]
    ).order_by('-fecha_movimiento')
    
    for transaccion in transacciones:
        row += 1
        ws[f'A{row}'] = transaccion.fecha_movimiento.strftime('%d/%m/%Y')
        ws[f'B{row}'] = transaccion.nombre
        ws[f'C{row}'] = transaccion.tipo.title()
        monto = transaccion.monto if transaccion.tipo == 'ingreso' else -transaccion.monto
        ws[f'D{row}'] = f"${monto:,.2f}"
        ws[f'E{row}'] = transaccion.id_cuenta.nombre
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"reporte_financiero_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

def exportar_csv(reporte, datos):
    """Exporta reporte a CSV"""
    import csv
    
    response = HttpResponse(content_type='text/csv')
    tipo_clean = reporte.get_tipo_reporte_display().replace(' ', '_').replace('/', '-')
    fecha_str = reporte.fecha_creacion.strftime('%Y%m%d_%H%M')
    filename = f"FinGest_Reporte_{tipo_clean}_{fecha_str}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Escribir encabezado
    writer.writerow(['Reporte:', reporte.nombre])
    writer.writerow(['Tipo:', reporte.get_tipo_reporte_display()])
    writer.writerow(['Fecha creación:', reporte.fecha_creacion.strftime('%Y-%m-%d %H:%M')])
    writer.writerow(['Período:', f"{reporte.fecha_inicio} - {reporte.fecha_fin}"])
    writer.writerow([])  # Línea vacía
    
    # Escribir datos según el tipo de reporte
    if reporte.tipo_reporte == 'ingresos_gastos':
        writer.writerow(['Fecha', 'Tipo', 'Descripción', 'Monto', 'Cuenta'])
        for item in datos.get('transacciones', []):
            writer.writerow([
                item.fecha_movimiento.strftime('%Y-%m-%d'),
                item.tipo.title(),
                item.nombre,
                f"${item.monto:.2f}",
                item.id_cuenta.nombre
            ])
    elif reporte.tipo_reporte == 'balance_cuentas':
        writer.writerow(['Cuenta', 'Saldo'])
        for item in datos.get('cuentas', []):
            writer.writerow([item.nombre, f"${item.saldo:.2f}"])
    elif reporte.tipo_reporte == 'gastos_categoria':
        writer.writerow(['Categoría', 'Monto', 'Porcentaje'])
        for item in datos.get('gastos_categoria', []):
            writer.writerow([item['nombre'], f"${item['total']:.2f}", f"{item['porcentaje']:.1f}%"])
    
    return response

def exportar_pdf_simple(request):
    """Exporta un reporte simple en PDF"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from io import BytesIO
    from datetime import datetime
    
    # Obtener fechas del filtro
    periodo = request.GET.get('periodo', 'mes_actual')
    fecha_inicio, fecha_fin = obtener_fechas_periodo(periodo)
    
    # Crear buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Center
    )
    
    # Título
    title = Paragraph("REPORTE FINANCIERO - FINGEST", title_style)
    story.append(title)
    
    # Información del período
    period_info = Paragraph(
        f"<b>Período:</b> {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}<br/>"
        f"<b>Generado:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal']
    )
    story.append(period_info)
    story.append(Spacer(1, 20))
    
    # Estadísticas generales
    stats = calcular_estadisticas_generales(request.user, fecha_inicio, fecha_fin)
    
    stats_data = [
        ['RESUMEN FINANCIERO', ''],
        ['Balance Total', f"${stats['balance_total']:,.2f}"],
        ['Ingresos del Período', f"${stats['total_ingresos']:,.2f}"],
        ['Gastos del Período', f"${stats['total_egresos']:,.2f}"],
        ['Ahorro Neto', f"${stats['ahorro_neto']:,.2f}"],
    ]
    
    stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1,  0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(stats_table)
    story.append(Spacer(1, 30))
    
    # Transacciones recientes
    transacciones = Movimiento.objects.filter(
        id_cuenta__id_usuario=request.user,
        fecha_movimiento__range=[fecha_inicio, fecha_fin]
    ).order_by('-fecha_movimiento')[:10]
    
    if transacciones:
        trans_title = Paragraph("<b>TRANSACCIONES RECIENTES</b>", styles['Heading2'])
        story.append(trans_title)
        
        trans_data = [['Fecha', 'Descripción', 'Tipo', 'Monto']]
        
        for transaccion in transacciones:
            trans_data.append([
                transaccion.fecha_movimiento.strftime('%d/%m/%Y'),
                transaccion.nombre[:30],
                transaccion.tipo.title(),
                f"${transaccion.monto:,.2f}"
            ])
        
        trans_table = Table(trans_data, colWidths=[1.2*inch, 2.5*inch, 1*inch, 1.3*inch])
        trans_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        story.append(trans_table)
    
    # Generar PDF
    doc.build(story)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f"reporte_financiero_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

def obtener_fechas_periodo(periodo):
    """Obtener fechas de inicio y fin según el período seleccionado"""
    from django.utils import timezone
    from datetime import datetime, timedelta
    import calendar
    
    now = timezone.now()
    
    if periodo == 'mes_actual':
        fecha_inicio = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        ultimo_dia = calendar.monthrange(now.year, now.month)[1]
        fecha_fin = now.replace(day=ultimo_dia, hour=23, minute=59, second=59)
    elif periodo == 'ultimo_mes':
        if now.month == 1:
            mes_anterior = now.replace(year=now.year-1, month=12, day=1)
        else:
            mes_anterior = now.replace(month=now.month-1, day=1)
        fecha_inicio = mes_anterior.replace(hour=0, minute=0, second=0, microsecond=0)
        ultimo_dia = calendar.monthrange(mes_anterior.year, mes_anterior.month)[1]
        fecha_fin = mes_anterior.replace(day=ultimo_dia, hour=23, minute=59, second=59)
    elif periodo == 'trimestre':
        fecha_inicio = (now - timedelta(days=90)).replace(hour=0, minute=0, second=0, microsecond=0)
        fecha_fin = now.replace(hour=23, minute=59, second=59)
    elif periodo == 'ano':
        fecha_inicio = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        fecha_fin = now.replace(hour=23, minute=59, second=59)
    else:
        # Por defecto mes actual
        fecha_inicio = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        ultimo_dia = calendar.monthrange(now.year, now.month)[1]
        fecha_fin = now.replace(day=ultimo_dia, hour=23, minute=59, second=59)
    
    return fecha_inicio, fecha_fin

# Funciones de exportación a Excel y PDF

@login_required
@fast_access_pin_verified
def exportar_excel(request):
    """Exportar datos financieros a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from datetime import datetime
    
    # Obtener fechas del filtro
    periodo = request.GET.get('periodo', 'mes_actual')
    fecha_inicio, fecha_fin = obtener_fechas_periodo(periodo)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Financiero"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Header del reporte
    ws['A1'] = "REPORTE FINANCIERO - FINGEST"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    ws['A2'].font = Font(bold=True)
    ws.merge_cells('A2:D2')
    
    ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws.merge_cells('A3:D3')
    
    # Estadísticas generales
    stats = calcular_estadisticas_generales(request.user, fecha_inicio, fecha_fin)
    
    row = 5
    ws[f'A{row}'] = "RESUMEN GENERAL"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws[f'A{row}'] = "Balance Total"
    ws[f'B{row}'] = f"${stats['balance_total']:,.2f}"
    
    row += 1
    ws[f'A{row}'] = "Ingresos del Período"
    ws[f'B{row}'] = f"${stats['total_ingresos']:,.2f}"
    
    row += 1
    ws[f'A{row}'] = "Gastos del Período"
    ws[f'B{row}'] = f"${stats['total_egresos']:,.2f}"
    
    row += 1
    ws[f'A{row}'] = "Ahorro Neto"
    ws[f'B{row}'] = f"${stats['ahorro_neto']:,.2f}"
    
    # Transacciones
    row += 3
    ws[f'A{row}'] = "TRANSACCIONES DETALLADAS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    headers = ['Fecha', 'Descripción', 'Tipo', 'Cuenta', 'Monto']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Obtener transacciones
    transacciones = Movimiento.objects.filter(
        id_cuenta__id_usuario=request.user,
        fecha_movimiento__range=[fecha_inicio, fecha_fin]
    ).order_by('-fecha_movimiento')
    
    for transaccion in transacciones:
        row += 1
        ws[f'A{row}'] = transaccion.fecha_movimiento.strftime('%d/%m/%Y')
        ws[f'B{row}'] = transaccion.nombre
        ws[f'C{row}'] = transaccion.tipo.title()
        ws[f'D{row}'] = transaccion.id_cuenta.nombre
        ws[f'E{row}'] = f"${transaccion.monto:,.2f}"
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"reporte_financiero_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@login_required
@fast_access_pin_verified
def exportar_pdf_simple(request):
    """Exportar reporte simple a PDF"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.utils import ImageReader
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from io import BytesIO
    from datetime import datetime
    
    # Obtener fechas del filtro
    periodo = request.GET.get('periodo', 'mes_actual')
    fecha_inicio, fecha_fin = obtener_fechas_periodo(periodo)
    
    # Crear buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Center
    )
    
    # Título
    title = Paragraph("REPORTE FINANCIERO - FINGEST", title_style)
    story.append(title)
    
    # Información del período
    period_info = Paragraph(
        f"<b>Período:</b> {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}<br/>"
        f"<b>Generado:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal']
    )
    story.append(period_info)
    story.append(Spacer(1, 20))
    
    # Estadísticas generales
    stats = calcular_estadisticas_generales(request.user, fecha_inicio, fecha_fin)
    
    stats_data = [
        ['RESUMEN FINANCIERO', ''],
        ['Balance Total', f"${stats['balance_total']:,.2f}"],
        ['Ingresos del Período', f"${stats['total_ingresos']:,.2f}"],
        ['Gastos del Período', f"${stats['total_egresos']:,.2f}"],
        ['Ahorro Neto', f"${stats['ahorro_neto']:,.2f}"],
    ]
    
    stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1,  0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(stats_table)
    story.append(Spacer(1, 30))
    
    # Transacciones recientes
    transacciones = Movimiento.objects.filter(
        id_cuenta__id_usuario=request.user,
        fecha_movimiento__range=[fecha_inicio, fecha_fin]
    ).order_by('-fecha_movimiento')[:10]
    
    if transacciones:
        trans_title = Paragraph("<b>TRANSACCIONES RECIENTES</b>", styles['Heading2'])
        story.append(trans_title)
        
        trans_data = [['Fecha', 'Descripción', 'Tipo', 'Monto']]
        
        for transaccion in transacciones:
            trans_data.append([
                transaccion.fecha_movimiento.strftime('%d/%m/%Y'),
                transaccion.nombre[:30],
                transaccion.tipo.title(),
                f"${transaccion.monto:,.2f}"
            ])
        
        trans_table = Table(trans_data, colWidths=[1.2*inch, 2.5*inch, 1*inch, 1.3*inch])
        trans_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        story.append(trans_table)
    
    # Generar PDF
    doc.build(story)
    
    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f"reporte_financiero_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

def procesar_datos_para_template(tipo_reporte, datos):
    """Procesa los datos del reporte para que sean más fáciles de usar en el template"""
    if not datos:
        return {}
    
    if tipo_reporte == 'gastos_categoria' and 'labels' in datos:
        # Crear lista de diccionarios para gastos por categoría
        total = sum(datos.get('data', []))
        resultado = []
        
        for i, label in enumerate(datos['labels']):
            monto = datos['data'][i] if i < len(datos['data']) else 0
            cantidad = datos.get('counts', [])[i] if i < len(datos.get('counts', [])) else 0
            porcentaje = (monto / total * 100) if total > 0 else 0
            
            resultado.append({
                'categoria': label,
                'monto': monto,
                'cantidad': cantidad,
                'porcentaje': porcentaje
            })
        
        return {
            'tipo': 'gastos_categoria',
            'items': resultado,
            'total': total,
            'total_cantidad': sum(datos.get('counts', []))
        }
    
    elif tipo_reporte == 'ingresos_egresos' and 'data' in datos:
        ingresos = datos['data'][0] if len(datos['data']) > 0 else 0
        egresos = datos['data'][1] if len(datos['data']) > 1 else 0
        balance = ingresos - egresos
        
        return {
            'tipo': 'ingresos_egresos',
            'ingresos': ingresos,
            'egresos': egresos,
            'balance': balance,
            'porcentaje_egresos': (egresos / ingresos * 100) if ingresos > 0 else 0,
            'porcentaje_balance': (balance / ingresos * 100) if ingresos > 0 else 0
        }
    
    elif tipo_reporte == 'subcuentas_analisis' and 'labels' in datos:
        resultado = []
        
        for i, label in enumerate(datos['labels']):
            saldo = datos['saldos'][i] if i < len(datos['saldos']) else 0
            cantidad = datos['cantidades'][i] if i < len(datos['cantidades']) else 0
            promedio = saldo / cantidad if cantidad > 0 else 0
            
            resultado.append({
                'tipo': label,
                'saldo': saldo,
                'cantidad': cantidad,
                'promedio': promedio
            })
        
        total_saldo = sum(datos['saldos'])
        total_cantidad = sum(datos['cantidades'])
        promedio_total = total_saldo / total_cantidad if total_cantidad > 0 else 0
        
        return {
            'tipo': 'subcuentas_analisis',
            'items': resultado,
            'total_saldo': total_saldo,
            'total_cantidad': total_cantidad,
            'promedio_total': promedio_total
        }
    
    elif tipo_reporte == 'flujo_efectivo' and 'labels' in datos:
        resultado = []
        
        for i, periodo in enumerate(datos['labels']):
            ingresos = datos['ingresos'][i] if i < len(datos['ingresos']) else 0
            egresos = datos['egresos'][i] if i < len(datos['egresos']) else 0
            balance = ingresos - egresos
            
            resultado.append({
                'periodo': periodo,
                'ingresos': ingresos,
                'egresos': egresos,
                'balance': balance
            })
        
        return {
            'tipo': 'flujo_efectivo',
            'items': resultado
        }
    
    return {'tipo': 'otros', 'datos_raw': datos}